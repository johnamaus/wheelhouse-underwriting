#!/usr/bin/env python3
"""
Wheelhouse Underwriting Tool — Local Server

Setup:
    pip install -r requirements.txt
    cp .env.example .env        # then edit with your real keys
    python server.py

Open: http://localhost:8000
"""
import os
import io
import json
import sqlite3
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_file, g, make_response
import re
import requests as req
from collections import Counter
import jwt
from werkzeug.security import generate_password_hash, check_password_hash

# ── Load .env if present ──
env_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

app = Flask(__name__)
BASE = "https://api.usewheelhouse.com/ss_api/v1/comp_set"
MR_BASE = "https://api.usewheelhouse.com/ss_api/v1/market_report"
DB_PATH = os.path.join("/tmp", "wh_underwriting.db")
JWT_SECRET = os.environ.get("JWT_SECRET", "change-me-in-production")
ADMIN_EMAILS = set(
    e.strip().lower()
    for e in os.environ.get("ADMIN_EMAILS", "").split(",")
    if e.strip()
)
# Seed emails — imported into DB on first init, then DB is source of truth
_SEED_EMAILS = set(
    e.strip().lower()
    for e in os.environ.get("ALLOWED_EMAILS", "").split(",")
    if e.strip()
)


# ── Auth helpers ──
def create_session_token(email, name):
    payload = {
        "email": email,
        "name": name,
        "exp": datetime.utcnow() + timedelta(days=7),
        "iat": datetime.utcnow(),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


def verify_session_token(token):
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
    except (jwt.ExpiredSignatureError, jwt.InvalidTokenError):
        return None


# ── SQLite helpers ──
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute("""
        CREATE TABLE IF NOT EXISTS searches (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at  TEXT NOT NULL,
            address     TEXT,
            lat         REAL,
            lng         REAL,
            radius      INTEGER,
            filters     TEXT,
            comp_count  INTEGER,
            results     TEXT,
            notes       TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            email         TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            name          TEXT NOT NULL DEFAULT '',
            is_admin      INTEGER NOT NULL DEFAULT 0,
            created_at    TEXT NOT NULL
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS allowed_emails (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            email      TEXT UNIQUE NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS reports (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at   TEXT NOT NULL,
            address      TEXT,
            lat          REAL,
            lng          REAL,
            radius       INTEGER,
            comp_count   INTEGER,
            snapshot     TEXT,
            preview_html TEXT,
            notes        TEXT
        )
    """)
    # Seed allowed emails from env var (only inserts new ones)
    for email in _SEED_EMAILS | ADMIN_EMAILS:
        db.execute(
            "INSERT OR IGNORE INTO allowed_emails (email, created_at) VALUES (?, ?)",
            (email, datetime.utcnow().isoformat()),
        )
    db.commit()
    db.close()


def get_headers():
    """Get API headers — prefer request headers, fall back to env vars."""
    return {
        "X-Integration-Api-Key": request.headers.get(
            "X-Integration-Api-Key",
            os.environ.get("WHEELHOUSE_INTEGRATION_KEY", ""),
        ),
        "X-User-API-Key": request.headers.get(
            "X-User-API-Key",
            os.environ.get("WHEELHOUSE_USER_KEY", ""),
        ),
    }


# ── Auth middleware ──
@app.before_request
def check_auth():
    # Allow auth routes and the main page without auth
    if request.path in ("/", "/auth/signup", "/auth/login", "/auth/me", "/auth/logout", "/auth/profile", "/auth/password"):
        return None
    # Protect all /api/* routes
    if request.path.startswith("/api/"):
        token = request.cookies.get("session_token")
        if not token:
            return jsonify({"error": "Not authenticated"}), 401
        user = verify_session_token(token)
        if not user:
            return jsonify({"error": "Invalid or expired session"}), 401
        g.user = user
    return None


# ── Auth routes ──
@app.route("/auth/signup", methods=["POST"])
def auth_signup():
    data = request.get_json()
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    name = (data.get("name") or "").strip()

    if not email or not password:
        return jsonify({"error": "Email and password are required"}), 400
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400

    db = get_db()
    allowed = db.execute("SELECT id FROM allowed_emails WHERE email = ?", (email,)).fetchone()
    if not allowed:
        return jsonify({"error": "This email is not authorized to access this tool"}), 403

    existing = db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
    if existing:
        return jsonify({"error": "An account with this email already exists"}), 409

    pw_hash = generate_password_hash(password)
    is_admin = 1 if email in ADMIN_EMAILS else 0
    db.execute(
        "INSERT INTO users (email, password_hash, name, is_admin, created_at) VALUES (?, ?, ?, ?, ?)",
        (email, pw_hash, name, is_admin, datetime.utcnow().isoformat()),
    )
    db.commit()

    token = create_session_token(email, name)
    resp = make_response(jsonify({"ok": True, "user": {"email": email, "name": name, "is_admin": bool(is_admin)}}))
    is_secure = request.url.startswith("https")
    resp.set_cookie("session_token", token, httponly=True, secure=is_secure,
                     samesite="Lax", max_age=7 * 86400)
    return resp


@app.route("/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json()
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""

    if not email or not password:
        return jsonify({"error": "Email and password are required"}), 400

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    if not user:
        return jsonify({"error": "Invalid email or password"}), 401

    if not check_password_hash(user["password_hash"], password):
        return jsonify({"error": "Invalid email or password"}), 401

    name = user["name"]
    # Auto-upgrade admin status if email is in ADMIN_EMAILS
    is_admin = bool(user["is_admin"])
    if email in ADMIN_EMAILS and not is_admin:
        db.execute("UPDATE users SET is_admin = 1 WHERE email = ?", (email,))
        db.commit()
        is_admin = True
    token = create_session_token(email, name)
    resp = make_response(jsonify({"ok": True, "user": {"email": email, "name": name, "is_admin": is_admin}}))
    is_secure = request.url.startswith("https")
    resp.set_cookie("session_token", token, httponly=True, secure=is_secure,
                     samesite="Lax", max_age=7 * 86400)
    return resp


@app.route("/auth/me")
def auth_me():
    token = request.cookies.get("session_token")
    if not token:
        return jsonify({"authenticated": False})
    user = verify_session_token(token)
    if not user:
        return jsonify({"authenticated": False})
    # Look up is_admin from DB
    db = get_db()
    db_user = db.execute("SELECT is_admin FROM users WHERE email = ?", (user["email"],)).fetchone()
    is_admin = bool(db_user and db_user["is_admin"])
    return jsonify({"authenticated": True, "user": {"email": user["email"], "name": user["name"], "is_admin": is_admin}})


@app.route("/auth/logout", methods=["POST"])
def auth_logout():
    resp = make_response(jsonify({"ok": True}))
    resp.delete_cookie("session_token")
    return resp


@app.route("/auth/profile", methods=["PUT"])
def auth_update_profile():
    token = request.cookies.get("session_token")
    if not token:
        return jsonify({"error": "Not authenticated"}), 401
    user = verify_session_token(token)
    if not user:
        return jsonify({"error": "Invalid session"}), 401

    data = request.get_json()
    name = (data.get("name") or "").strip()
    new_email = (data.get("email") or "").strip().lower()
    current_email = user["email"]

    db = get_db()
    if new_email and new_email != current_email:
        # Check if new email is already taken
        existing = db.execute("SELECT id FROM users WHERE email = ?", (new_email,)).fetchone()
        if existing:
            return jsonify({"error": "Email already in use"}), 409
        db.execute("UPDATE users SET email = ?, name = ? WHERE email = ?", (new_email, name, current_email))
        # Also update allowed_emails table
        db.execute("UPDATE allowed_emails SET email = ? WHERE email = ?", (new_email, current_email))
    else:
        new_email = current_email
        db.execute("UPDATE users SET name = ? WHERE email = ?", (name, current_email))
    db.commit()

    # Issue new token with updated info
    token = create_session_token(new_email, name)
    db_user = db.execute("SELECT is_admin FROM users WHERE email = ?", (new_email,)).fetchone()
    is_admin = bool(db_user and db_user["is_admin"])
    resp = make_response(jsonify({"ok": True, "user": {"email": new_email, "name": name, "is_admin": is_admin}}))
    is_secure = request.url.startswith("https")
    resp.set_cookie("session_token", token, httponly=True, secure=is_secure,
                     samesite="Lax", max_age=7 * 86400)
    return resp


@app.route("/auth/password", methods=["PUT"])
def auth_change_password():
    token = request.cookies.get("session_token")
    if not token:
        return jsonify({"error": "Not authenticated"}), 401
    user = verify_session_token(token)
    if not user:
        return jsonify({"error": "Invalid session"}), 401

    data = request.get_json()
    current_pw = data.get("current_password") or ""
    new_pw = data.get("new_password") or ""

    if len(new_pw) < 6:
        return jsonify({"error": "New password must be at least 6 characters"}), 400

    db = get_db()
    db_user = db.execute("SELECT password_hash FROM users WHERE email = ?", (user["email"],)).fetchone()
    if not db_user or not check_password_hash(db_user["password_hash"], current_pw):
        return jsonify({"error": "Current password is incorrect"}), 401

    new_hash = generate_password_hash(new_pw)
    db.execute("UPDATE users SET password_hash = ? WHERE email = ?", (new_hash, user["email"]))
    db.commit()
    return jsonify({"ok": True})


# ── Admin routes ──
def require_admin():
    """Check that the current user is an admin. Returns error response or None."""
    token = request.cookies.get("session_token")
    if not token:
        return jsonify({"error": "Not authenticated"}), 401
    user = verify_session_token(token)
    if not user:
        return jsonify({"error": "Invalid session"}), 401
    db = get_db()
    db_user = db.execute("SELECT is_admin FROM users WHERE email = ?", (user["email"],)).fetchone()
    if not db_user or not db_user["is_admin"]:
        return jsonify({"error": "Admin access required"}), 403
    return None


@app.route("/api/admin/allowed-emails")
def admin_list_emails():
    err = require_admin()
    if err:
        return err
    db = get_db()
    rows = db.execute("""
        SELECT a.email, a.created_at,
               COALESCE(u.is_admin, 0) as is_admin,
               CASE WHEN u.id IS NOT NULL THEN 1 ELSE 0 END as has_account
        FROM allowed_emails a
        LEFT JOIN users u ON a.email = u.email
        ORDER BY a.created_at DESC
    """).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/allowed-emails", methods=["POST"])
def admin_add_email():
    err = require_admin()
    if err:
        return err
    data = request.get_json()
    email = (data.get("email") or "").strip().lower()
    if not email or "@" not in email:
        return jsonify({"error": "Valid email is required"}), 400
    db = get_db()
    is_admin = data.get("is_admin", False)
    existing = db.execute("SELECT id FROM allowed_emails WHERE email = ?", (email,)).fetchone()
    if existing:
        return jsonify({"error": "Email already in the allowed list"}), 409
    db.execute("INSERT INTO allowed_emails (email, created_at) VALUES (?, ?)",
               (email, datetime.utcnow().isoformat()))
    # If marking as admin and user already has an account, set their admin flag
    if is_admin:
        db.execute("UPDATE users SET is_admin = 1 WHERE email = ?", (email,))
        ADMIN_EMAILS.add(email)
    db.commit()
    return jsonify({"ok": True, "email": email})


@app.route("/api/admin/allowed-emails/<path:email>", methods=["DELETE"])
def admin_remove_email(email):
    err = require_admin()
    if err:
        return err
    email = email.strip().lower()
    db = get_db()
    db.execute("DELETE FROM allowed_emails WHERE email = ?", (email,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/toggle-admin/<path:email>", methods=["POST"])
def admin_toggle_admin(email):
    err = require_admin()
    if err:
        return err
    email = email.strip().lower()
    db = get_db()
    user = db.execute("SELECT is_admin FROM users WHERE email = ?", (email,)).fetchone()
    if not user:
        return jsonify({"error": "User has not signed up yet"}), 404
    new_val = 0 if user["is_admin"] else 1
    db.execute("UPDATE users SET is_admin = ? WHERE email = ?", (new_val, email))
    db.commit()
    return jsonify({"ok": True, "is_admin": bool(new_val)})


# ── Page routes ──
@app.route("/")
def index():
    return send_file(os.path.join(os.path.dirname(__file__), "index.html"))


# ── API proxy routes ──
def _try_nominatim(address):
    """Try OpenStreetMap Nominatim geocoder."""
    r = req.get(
        "https://nominatim.openstreetmap.org/search",
        params={"format": "json", "q": address, "limit": 5, "addressdetails": 1},
        headers={"User-Agent": "WheelhouseUnderwriting/1.0 (john@usewheelhouse.com)"},
        timeout=10,
    )
    r.raise_for_status()
    return r.json()


def _try_census(address):
    """Fallback: US Census Bureau geocoder (excellent US address coverage)."""
    r = req.get(
        "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress",
        params={"address": address, "benchmark": "Public_AR_Current", "format": "json"},
        timeout=10,
    )
    r.raise_for_status()
    matches = r.json().get("result", {}).get("addressMatches", [])
    if not matches:
        return []
    # Convert to Nominatim-compatible format
    return [
        {
            "lat": str(m["coordinates"]["y"]),
            "lon": str(m["coordinates"]["x"]),
            "display_name": m.get("matchedAddress", address),
        }
        for m in matches
    ]


@app.route("/api/geocode")
def geocode():
    address = request.args.get("q", "")
    try:
        # 1) Try Nominatim
        data = _try_nominatim(address)
        if data:
            return jsonify(data)
        # 2) Retry Nominatim with simplified query
        parts = [p.strip() for p in address.split(",")]
        if len(parts) >= 2:
            data = _try_nominatim(", ".join(parts[:2]))
            if data:
                return jsonify(data)
        # 3) Fallback to US Census geocoder
        print(f"  ℹ Nominatim miss — trying Census geocoder for: {address}")
        census = _try_census(address)
        if census:
            return jsonify(census)
        return jsonify([])
    except Exception as e:
        print(f"  ⚠ Geocode error: {e}")
        return jsonify({"error": str(e)}), 502


@app.route("/api/brand-colors")
def brand_colors():
    """Extract brand/accent colors from a website's HTML/CSS."""
    url = request.args.get("url", "").strip()
    if not url:
        return jsonify({"error": "No URL provided"}), 400
    if not url.startswith("http"):
        url = "https://" + url
    try:
        r = req.get(url, timeout=8, headers={
            "User-Agent": "Mozilla/5.0 (compatible; WheelhouseBot/1.0)"
        })
        html = r.text[:200000]  # limit to 200k chars

        # Collect hex colors from inline styles and <style> blocks
        hex6 = re.findall(r'#([0-9a-fA-F]{6})\b', html)
        hex3 = re.findall(r'#([0-9a-fA-F]{3})\b', html)
        # Expand 3-char hex to 6-char
        all_hex = [h.lower() for h in hex6]
        all_hex += [f"{h[0]}{h[0]}{h[1]}{h[1]}{h[2]}{h[2]}".lower() for h in hex3]

        # Also grab rgb()/rgba()
        rgb_matches = re.findall(r'rgb\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)', html)
        for r_, g_, b_ in rgb_matches:
            all_hex.append(f"{int(r_):02x}{int(g_):02x}{int(b_):02x}")

        # Filter out near-white, near-black, and pure grays
        def is_brand(h):
            r_, g_, b_ = int(h[:2], 16), int(h[2:4], 16), int(h[4:6], 16)
            brightness = (r_ * 299 + g_ * 587 + b_ * 114) / 1000
            if brightness > 240 or brightness < 20:
                return False
            # filter pure grays (low saturation)
            mx, mn = max(r_, g_, b_), min(r_, g_, b_)
            if mx - mn < 25:
                return False
            return True

        filtered = [h for h in all_hex if is_brand(h)]
        counts = Counter(filtered)
        top = counts.most_common(5)  # top 5 most-used brand colors
        colors = [{"hex": f"#{h}", "count": c,
                    "rgb": [int(h[:2], 16), int(h[2:4], 16), int(h[4:6], 16)]}
                   for h, c in top]
        return jsonify({"url": url, "colors": colors})
    except Exception as e:
        print(f"  ⚠ Brand color extraction error: {e}")
        return jsonify({"error": str(e), "colors": []}), 200


@app.route("/api/autocomplete")
def autocomplete():
    """Fast address suggestions — ArcGIS suggest (single request, no coord resolution)."""
    q = request.args.get("q", "").strip()
    if len(q) < 3:
        return jsonify([])
    try:
        r = req.get(
            "https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/suggest",
            params={"text": q, "f": "json", "countryCode": "USA", "maxSuggestions": 6},
            timeout=3,
        )
        if not r.ok:
            return jsonify([])
        suggestions = r.json().get("suggestions", [])
        return jsonify([{"display_name": s.get("text", ""), "magicKey": s.get("magicKey", "")}
                        for s in suggestions if s.get("text")])
    except Exception as e:
        print(f"  Autocomplete error: {e}")
        return jsonify([])


@app.route("/api/resolve-address")
def resolve_address():
    """Resolve a selected suggestion to lat/lon using its magicKey."""
    text = request.args.get("text", "").strip()
    magic = request.args.get("magicKey", "").strip()
    if not text:
        return jsonify({"error": "missing text"}), 400
    try:
        params = {"SingleLine": text, "f": "json", "maxLocations": 1, "countryCode": "USA"}
        if magic:
            params["magicKey"] = magic
        r = req.get(
            "https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/findAddressCandidates",
            params=params, timeout=3,
        )
        if r.ok:
            cands = r.json().get("candidates", [])
            if cands:
                loc = cands[0].get("location", {})
                return jsonify({
                    "display_name": cands[0].get("address", text),
                    "lat": str(loc.get("y", "")),
                    "lon": str(loc.get("x", "")),
                })
        return jsonify({"error": "no results"}), 404
    except Exception as e:
        print(f"  Resolve error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/market-report/markets")
def market_report_markets():
    """List available markets for a country (Market Report API)."""
    headers = {"X-Integration-Api-Key": os.environ.get("WHEELHOUSE_MARKET_REPORT_KEY", "")}
    params = {k: v for k, v in request.args.items()}
    if "country_code" not in params:
        params["country_code"] = "US"
    try:
        r = req.get(f"{MR_BASE}/", headers=headers, params=params, timeout=10)
        return (r.text, r.status_code, {"Content-Type": "application/json"})
    except Exception as e:
        print(f"  ⚠ Market report markets error: {e}")
        return jsonify({"error": str(e)}), 502


@app.route("/api/market-report/<market_id>/time-series")
def market_report_time_series(market_id):
    """Fetch daily time-series data for a market (Market Report API).

    Metrics: occupancy, available_rooms, asking_rate, nightly_rate, revpar, revenue, lead_time
    Filters: start_date, end_date, bedrooms (0,1,2,3,4+), performance (low,average,high), property_type
    """
    headers = {"X-Integration-Api-Key": os.environ.get("WHEELHOUSE_MARKET_REPORT_KEY", "")}
    # Use to_dict(flat=False) to preserve multi-value params like metric=occupancy&metric=nightly_rate
    params = []
    for k, values in request.args.lists():
        for v in values:
            params.append((k, v))
    try:
        print(f"  [MR] time-series market={market_id} params={params}")
        r = req.get(f"{MR_BASE}/{market_id}/time_series", headers=headers, params=params, timeout=15)
        print(f"  [MR] time-series response: {r.status_code} {r.text[:300]}")
        return (r.text, r.status_code, {"Content-Type": "application/json"})
    except Exception as e:
        print(f"  ⚠ Market report time-series error: {e}")
        return jsonify({"error": str(e)}), 502


@app.route("/api/market-report/<market_id>/distribution")
def market_report_distribution(market_id):
    """Fetch monthly metric distributions for a market (Market Report API)."""
    headers = {"X-Integration-Api-Key": os.environ.get("WHEELHOUSE_MARKET_REPORT_KEY", "")}
    params = {k: v for k, v in request.args.items()}
    try:
        r = req.get(f"{MR_BASE}/{market_id}/distribution", headers=headers, params=params, timeout=15)
        return (r.text, r.status_code, {"Content-Type": "application/json"})
    except Exception as e:
        print(f"  ⚠ Market report distribution error: {e}")
        return jsonify({"error": str(e)}), 502


@app.route("/api/candidates")
def candidates():
    headers = get_headers()
    params = {k: v for k, v in request.args.items()}
    r = req.get(f"{BASE}/candidates", headers=headers, params=params)
    # Debug: log source fields from API response
    try:
        data = r.json()
        items = data if isinstance(data, list) else data.get("listings") or data.get("results") or data.get("candidates") or []
        if items and len(items) > 0:
            sample = items[0]
            sources = set(x.get("source") for x in items if "source" in x)
            print(f"  [DEBUG] {len(items)} comps returned. Sources: {sources}")
            print(f"  [DEBUG] Sample comp keys: {list(sample.keys())[:20]}")
    except Exception as e:
        print(f"  [DEBUG] Could not parse response: {e}")
    return (r.text, r.status_code, {"Content-Type": "application/json"})


@app.route("/api/candidates/listing/<listing_id>")
def listing(listing_id):
    headers = get_headers()
    r = req.get(f"{BASE}/candidates/listing/{listing_id}", headers=headers)
    return (r.text, r.status_code, {"Content-Type": "application/json"})


@app.route("/api/candidates/listings")
def listings():
    headers = get_headers()
    params = {k: v for k, v in request.args.items()}
    r = req.get(f"{BASE}/candidates/listings", headers=headers, params=params)
    return (r.text, r.status_code, {"Content-Type": "application/json"})


@app.route("/api/export-comps", methods=["POST"])
def export_comps():
    """Export comp data as Excel — receives full comp objects from the client."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    body = request.get_json(force=True)
    all_comps = body.get("comps", [])
    if not all_comps:
        return jsonify({"error": "No comps provided"}), 400

    # Collect all unique keys across comps
    all_keys = []
    seen = set()
    for comp in all_comps:
        for k in comp.keys():
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comp Set Data"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # Write headers
    for col_idx, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    alt_fill = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
    for row_idx, comp in enumerate(all_comps, 2):
        for col_idx, key in enumerate(all_keys, 1):
            val = comp.get(key, "")
            # Flatten lists/dicts to string
            if isinstance(val, (list, dict)):
                val = json.dumps(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-width columns (capped at 40)
    for col_idx, key in enumerate(all_keys, 1):
        max_len = len(str(key))
        for row_idx in range(2, len(all_comps) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 3

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="comp_set_export.xlsx")


# ── Search history routes ──
@app.route("/api/searches", methods=["GET"])
def list_searches():
    db = get_db()
    rows = db.execute(
        "SELECT id, created_at, address, lat, lng, radius, filters, comp_count, notes "
        "FROM searches ORDER BY created_at DESC LIMIT 100"
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/searches", methods=["POST"])
def save_search():
    data = request.get_json()
    db = get_db()
    db.execute(
        "INSERT INTO searches (created_at, address, lat, lng, radius, filters, comp_count, results, notes) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            datetime.utcnow().isoformat(),
            data.get("address"),
            data.get("lat"),
            data.get("lng"),
            data.get("radius"),
            json.dumps(data.get("filters", {})),
            data.get("comp_count", 0),
            json.dumps(data.get("results", [])),
            data.get("notes", ""),
        ),
    )
    db.commit()
    return jsonify({"ok": True, "id": db.execute("SELECT last_insert_rowid()").fetchone()[0]})


@app.route("/api/searches/<int:search_id>")
def get_search(search_id):
    db = get_db()
    row = db.execute("SELECT * FROM searches WHERE id = ?", (search_id,)).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    result = dict(row)
    result["results"] = json.loads(result["results"]) if result["results"] else []
    result["filters"] = json.loads(result["filters"]) if result["filters"] else {}
    return jsonify(result)


@app.route("/api/searches/<int:search_id>", methods=["DELETE"])
def delete_search(search_id):
    db = get_db()
    db.execute("DELETE FROM searches WHERE id = ?", (search_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/searches/<int:search_id>/notes", methods=["PUT"])
def update_notes(search_id):
    data = request.get_json()
    db = get_db()
    db.execute("UPDATE searches SET notes = ? WHERE id = ?", (data.get("notes", ""), search_id))
    db.commit()
    return jsonify({"ok": True})


# ── Report History ──────────────────────────────────────────────
@app.route("/api/reports", methods=["GET"])
def list_reports():
    db = get_db()
    rows = db.execute(
        "SELECT id, created_at, address, lat, lng, comp_count, notes "
        "FROM reports ORDER BY created_at DESC LIMIT 100"
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/reports", methods=["POST"])
def save_report():
    data = request.get_json()
    db = get_db()
    db.execute(
        "INSERT INTO reports (created_at, address, lat, lng, radius, comp_count, snapshot, preview_html, notes) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            datetime.utcnow().isoformat(),
            data.get("address"),
            data.get("lat"),
            data.get("lng"),
            data.get("radius"),
            data.get("comp_count", 0),
            json.dumps(data.get("snapshot", {})),
            data.get("preview_html", ""),
            data.get("notes", ""),
        ),
    )
    db.commit()
    return jsonify({"ok": True, "id": db.execute("SELECT last_insert_rowid()").fetchone()[0]})


@app.route("/api/reports/<int:report_id>")
def get_report(report_id):
    db = get_db()
    row = db.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    result = dict(row)
    result["snapshot"] = json.loads(result["snapshot"]) if result["snapshot"] else {}
    return jsonify(result)


@app.route("/api/reports/<int:report_id>", methods=["DELETE"])
def delete_report(report_id):
    db = get_db()
    db.execute("DELETE FROM reports WHERE id = ?", (report_id,))
    db.commit()
    return jsonify({"ok": True})


# Always init DB (needed for Vercel where __main__ doesn't run)
init_db()

if __name__ == "__main__":
    print()
    print("  🏠 Wheelhouse Underwriting Tool")
    print("  ─────────────────────────────────")
    print("  Open http://localhost:8000")
    print(f"  Database: {DB_PATH}")
    print()
    app.run(host="0.0.0.0", port=8000, debug=True)
